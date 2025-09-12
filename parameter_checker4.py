#!/usr/bin/env python3
"""
参数核查系统 - 重新设计版本
支持双分表结构、复杂条件表达式、嵌套验证规则
Parameter Checker System - Redesigned Version
Supports dual-table structure, complex condition expressions, nested validation rules
"""

import pandas as pd
import logging
import re
from typing import Dict, List, Any, Optional, Tuple, Set, Callable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('parameter_checker.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)


def get_group_keys(df):
    return set(df.groupby(['f_site_id', 'f_cell_id']).groups.keys())


class ParameterChecker:
    """
    参数核查器类 - 支持双分表结构和复杂嵌套验证
    """
    # 类级常量：支持的运算符（按长度排序，确保长运算符优先匹配）
    OPERATORS = ['>=', '<=', '!=', '>', '<', '=']
    # 逻辑运算符正则（用于匹配完整单词）
    LOGICAL_OP_PATTERN = re.compile(r'\band\b|\bor\b', re.IGNORECASE)

    def __init__(self, knowledge_file="参数知识库.xlsx"):
        """初始化参数核查器"""
        self.knowledge_file = knowledge_file
        self.parameter_info: Dict[str, Dict[str, Any]] = {}  # 参数信息表
        self.validation_rules: Dict[str, Dict[str, Any]] = {}  # 验证规则表

        # 加载知识库
        self.load_knowledge_base(knowledge_file)

    def load_knowledge_base(self, file_path: str) -> bool:
        """加载双分表知识库"""
        try:
            # 加载参数信息表
            param_success = self.load_parameter_info(file_path, "参数信息")
            # 加载验证规则表
            rule_success = self.load_validation_rules(file_path, "验证规则")

            if param_success and rule_success:
                logger.info(f"知识库加载成功: {len(self.parameter_info)}个参数, {len(self.validation_rules)}个验证规则")
                return True
            else:
                logger.warning("知识库加载部分失败")
                return False

        except FileNotFoundError:
            logger.info(f"文件 {file_path} 不存在，正在生成示例文件...")
            # 生成示例文件
            self.create_sample_excel()
            # 重新加载
            try:
                param_success = self.load_parameter_info(file_path, "参数信息")
                rule_success = self.load_validation_rules(file_path, "验证规则")
                if param_success and rule_success:
                    logger.info(
                        f"示例知识库加载成功: {len(self.parameter_info)}个参数, {len(self.validation_rules)}个验证规则")
                    return True
            except Exception as e:
                logger.error(f"重新加载知识库失败: {str(e)}")
            return False
        except Exception as e:
            logger.error(f"加载知识库失败: {str(e)}")
            return False

    def load_parameter_info(self, file_path: str, sheet_name: str) -> bool:
        """加载参数信息表"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)

            # 验证必要列
            required_columns = ['MO名称', 'MO描述', '场景类型', '参数名称', '参数ID', '参数类型', '参数含义', '值描述']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                logger.error(f"参数信息表缺少必要列: {missing_columns}")
                return False

            self.parameter_info = {}

            # 按MO名称和参数名称分组
            for (mo_name, param_name), group in df.groupby(['MO名称', '参数名称'], dropna=False):
                # 使用第一行的信息
                row = group.iloc[0]

                # 初始化MO信息
                if mo_name not in self.parameter_info:
                    self.parameter_info[mo_name] = {
                        'mo_description': str(row.get('MO描述', '')).strip(),
                        'scenario': str(row.get('场景类型', '')).strip(),
                        'parameters': {}
                    }

                # 添加参数信息
                param_type = str(row.get('参数类型', 'single')).strip()
                self.parameter_info[mo_name]['parameters'][param_name] = {
                    'parameter_id': str(row.get('参数ID', '')).strip(),
                    'parameter_type': param_type,
                    'parameter_description': str(row.get('参数含义', '')).strip(),
                    'value_description': str(row.get('值描述', '')).strip()
                }

            logger.info(f"参数信息表加载成功: {len(self.parameter_info)} 个MO")
            return True

        except FileNotFoundError:
            raise  # 重新抛出FileNotFoundError让上层处理
        except Exception as e:
            logger.error(f"加载参数信息表失败: {str(e)}")
            return False

    def load_validation_rules(self, file_path: str, sheet_name: str) -> bool:
        """加载验证规则表"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)

            # 验证必要列
            required_columns = ['校验ID', '校验类型', 'MO名称', '条件表达式', '期望值表达式', '错误描述', '继续校验ID']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                logger.error(f"验证规则表缺少必要列: {missing_columns}")
                return False

            self.validation_rules = {}

            for _, row in df.iterrows():
                rule_id = str(row.get('校验ID', '')).strip()
                if not rule_id or rule_id == 'nan':
                    continue

                self.validation_rules[rule_id] = {
                    'rule_id': rule_id,
                    'check_type': str(row.get('校验类型', '')).strip(),
                    'mo_name': str(row.get('MO名称', '')).strip(),
                    'condition_expression': str(row.get('条件表达式', '')).strip(),
                    'expected_expression': str(row.get('期望值表达式', '')).strip(),
                    'error_description': str(row.get('错误描述', '')).strip(),
                    'next_check_id': str(row.get('继续校验ID', '')).strip() if str(
                        row.get('继续校验ID', '')).strip() != 'nan' else None
                }

            logger.info(f"验证规则表加载成功: {len(self.validation_rules)} 个规则")
            return True

        except FileNotFoundError:
            raise  # 重新抛出FileNotFoundError让上层处理
        except Exception as e:
            logger.error(f"加载验证规则表失败: {str(e)}")
            return False

    def parse_condition_expression(self, expression: str, data_row: Dict[str, Any]) -> bool:
        """
        解析复杂条件表达式
        支持格式: （参数名1=值1and参数名2=值2）or（参数名3>值3and参数名2!=值2）
        """
        if not expression or expression == 'nan':
            return True

        try:
            # 标准化表达式：处理中文括号和操作符
            expression = self._normalize_condition_expression(expression)

            # 解析单个条件
            def evaluate_single_condition(cond: str) -> bool:
                cond = cond.strip()

                # 支持的运算符（按长度降序排列避免匹配问题）
                operators = ['>=', '<=', '!=', '>', '<', '=']

                for op in operators:
                    if op in cond:
                        parts = cond.split(op, 1)
                        if len(parts) == 2:
                            param_name = parts[0].strip()
                            expected_value = parts[1].strip()

                            # 获取实际值
                            actual_value = str(data_row.get(param_name, '')).strip()

                            # 执行比较
                            return self._compare_values(actual_value, op, expected_value)

                logger.warning(f"无法解析条件: {cond}")
                return False

            # 递归处理括号和逻辑运算符
            def evaluate_expression(expr: str) -> bool:
                expr = expr.strip()

                # 处理最内层括号
                while '(' in expr and ')' in expr:
                    start = expr.rfind('(')
                    end = expr.find(')', start)
                    if end == -1:
                        break

                    # 评估括号内的表达式
                    inner_expr = expr[start + 1:end]
                    result = self._evaluate_simple_expression(inner_expr, evaluate_single_condition)

                    # 替换括号及其内容为结果
                    expr = expr[:start] + str(result).lower() + expr[end + 1:]

                # 评估剩余表达式
                return self._evaluate_simple_expression(expr, evaluate_single_condition)

            return evaluate_expression(expression)

        except Exception as e:
            logger.error(f"解析条件表达式失败: {expression}, 错误: {str(e)}")
            return False

    def _normalize_condition_expression(self, expression: str) -> str:
        """标准化条件表达式：处理中文括号和操作符，表达式正则，and or前后加空格，' and '"""
        import re

        # 将中文括号转换为英文括号
        expression = expression.replace('（', '(').replace('）', ')')

        # 在逻辑操作符前后添加空格，处理无空格的情况
        expression = re.sub(r'(?<!\s)and(?!\s)', ' and ', expression)
        expression = re.sub(r'(?<!\s)or(?!\s)', ' or ', expression)

        # 处理中文逻辑操作符
        expression = expression.replace('且', ' and ').replace('或', ' or ')

        # 清理多余空格
        expression = re.sub(r'\s+', ' ', expression)

        return expression.strip()

    def _evaluate_simple_expression(self, expr: str, eval_func) -> bool:
        """评估简单表达式（不含括号）"""
        expr = expr.strip()

        # 处理 or 运算符
        if ' or ' in expr:
            parts = expr.split(' or ')
            return any(self._evaluate_simple_expression(part.strip(), eval_func) for part in parts)

        # 处理 and 运算符
        if ' and ' in expr:
            parts = expr.split(' and ')
            return all(self._evaluate_simple_expression(part.strip(), eval_func) for part in parts)

        # 处理布尔值
        if expr.lower() == 'true':
            return True
        elif expr.lower() == 'false':
            return False

        # 单个条件
        return eval_func(expr)

    def _compare_values(self, actual: str, operator: str, expected: str) -> bool:
        """比较两个值"""
        try:
            # 尝试数值比较
            try:
                actual_num = float(actual)
                expected_num = float(expected)

                if operator == '=':
                    return actual_num == expected_num
                elif operator == '!=':
                    return actual_num != expected_num
                elif operator == '>':
                    return actual_num > expected_num
                elif operator == '<':
                    return actual_num < expected_num
                elif operator == '>=':
                    return actual_num >= expected_num
                elif operator == '<=':
                    return actual_num <= expected_num
            except ValueError:
                # 字符串比较
                if operator == '=':
                    return actual == expected
                elif operator == '!=':
                    return actual != expected
                else:
                    # 字符串不支持大小比较
                    logger.warning(f"字符串不支持运算符 {operator}: {actual} {operator} {expected}")
                    return False

        except Exception as e:
            logger.error(f"值比较失败: {actual} {operator} {expected}, 错误: {str(e)}")
            return False

        return False

    def _is_complex_expression(self, expression: str) -> bool:
        """判断是否为复杂表达式（包含括号或逻辑运算符）"""
        has_parentheses = '(' in expression and ')' in expression
        has_logical_ops = self.LOGICAL_OP_PATTERN.search(expression) is not None
        return has_parentheses or has_logical_ops

    def _is_complex_expression_improved(self, expression: str) -> bool:
        """改进的复杂表达式判断 - 解决Unicode边界和无空格问题"""
        has_parentheses = '(' in expression and ')' in expression
        # 更宽泛的逻辑运算符检测，处理标准化后的表达式
        expr_lower = expression.lower()
        has_logical_ops = (' and ' in expr_lower) or (' or ' in expr_lower)
        return has_parentheses or has_logical_ops

    def _parse_complex_expression(self, expression: str) -> Dict[str, Any]:
        """解析复杂表达式 - 使用修复版解析器"""
        try:
            # 直接使用修复版的参数提取方法
            params_info = self.new_extract_param_details_fixed(expression)

            return {
                'type': 'complex',
                'expression': expression,
                'params': params_info
            }
        except Exception as e:
            # 复杂表达式解析失败时返回原始表达式和错误信息
            logger.error(f"复杂表达式解析失败: {str(e)}")
            return {
                'type': 'complex',
                'expression': expression,
                'params': [],
                'error': f"解析复杂表达式失败: {str(e)}"
            }

    def _process_parentheses(self, expression: str) -> str:
        """
        标准化表达式但保持括号结构，避免过早删除括号
        只进行基本的标准化处理，保持深度感知解析的有效性
        """
        # 只进行基本标准化，不删除括号
        normalized_expr = self._normalize_condition_expression(expression)

        # 保持原始括号结构，让后续的参数提取过程处理嵌套逻辑
        return normalized_expr

    def _extract_param_details(self, expression: str) -> List[Dict[str, Any]]:
        """
        从表达式中提取参数详细信息 - 使用修复版解析器
        正确处理参数名中的括号，统一处理单值/多值参数
        """
        try:
            # 使用修复版的解析策略
            return self.new_extract_param_details_fixed(expression)
        except Exception as e:
            logger.error(f"修复版参数提取失败，使用备用方法: {str(e)}")
            # 回退到简化的备用方案
            return self._extract_param_details_fallback(expression)

    def _extract_param_details_improved(self, expression: str) -> List[Dict[str, Any]]:
        """
        改进的参数详细信息提取 - 使用重写的安全解析器
        注意：此方法已弃用，现在直接使用修复版解析器避免括号分割问题
        """
        logger.warning("_extract_param_details_improved已弃用，使用安全解析器")
        # 直接使用重写的安全解析器，避免括号分割问题
        return self.new_extract_param_details_fixed(expression)

    def _extract_param_details_fallback(self, expression: str) -> List[Dict[str, Any]]:
        """
        回退的参数提取方法（原有逻辑的简化版）
        """
        params_info = []
        working_expr = expression.strip()

        while working_expr:
            found = False
            for op in self.OPERATORS:
                op_index = working_expr.find(op)
                if op_index != -1:
                    # 提取参数名
                    param_name = working_expr[:op_index].strip()
                    if not param_name:
                        working_expr = working_expr[op_index + len(op):].strip()
                        continue

                    # 提取参数值和更新工作表达式
                    param_value, working_expr = self._extract_param_value(
                        working_expr, op_index, len(op))

                    # 解析参数详情（单值/多值）
                    param_detail = self._parse_param_detail(
                        param_name, param_value, op)
                    params_info.append(param_detail)

                    found = True
                    break

            if not found:
                break

        return params_info

    def _extract_param_value(self, expr: str, op_index: int, op_length: int) -> Tuple[str, str]:
        """提取参数值并计算下一个处理位置 - 支持嵌套括号和复杂逻辑"""
        remaining_expr = expr[op_index + op_length:].strip()

        if not remaining_expr:
            return "", ""

        # 如果以括号开始，需要找到匹配的右括号
        if remaining_expr.startswith('('):
            return self._extract_parenthesized_value(remaining_expr)

        # 处理非括号情况，考虑括号嵌套
        paren_depth = 0
        i = 0

        while i < len(remaining_expr):
            char = remaining_expr[i]

            if char == '(':
                paren_depth += 1
            elif char == ')':
                paren_depth -= 1
            elif paren_depth == 0:  # 只在括号外查找逻辑运算符
                # 检查是否为逻辑运算符
                if self._is_logical_operator_at_position(remaining_expr, i):
                    param_value = remaining_expr[:i].strip()
                    # 跳过逻辑运算符
                    op_len = self._get_logical_operator_length(remaining_expr, i)
                    next_working_expr = remaining_expr[i + op_len:].strip()
                    return param_value, next_working_expr

            i += 1

        # 没有找到逻辑运算符，整个剩余部分就是值
        return remaining_expr.strip(), ""

    def _extract_parenthesized_value(self, expr: str) -> Tuple[str, str]:
        """提取括号包围的值"""
        if not expr.startswith('('):
            return expr, ""

        paren_depth = 0
        i = 0

        for i, char in enumerate(expr):
            if char == '(':
                paren_depth += 1
            elif char == ')':
                paren_depth -= 1
                if paren_depth == 0:
                    # 找到匹配的右括号
                    value_with_parens = expr[:i + 1]
                    remaining = expr[i + 1:].strip()

                    # 检查后面是否有逻辑运算符
                    if remaining and self._is_logical_operator_at_position(remaining, 0):
                        op_len = self._get_logical_operator_length(remaining, 0)
                        remaining = remaining[op_len:].strip()

                    return value_with_parens, remaining

        # 括号不匹配的情况
        return expr, ""

    def _is_logical_operator_at_position(self, expr: str, pos: int) -> bool:
        """检查指定位置是否为逻辑运算符"""
        if pos >= len(expr):
            return False

        # 检查 'and'
        if (pos + 3 <= len(expr) and
                expr[pos:pos + 3].lower() == 'and' and
                (pos == 0 or not expr[pos - 1].isalnum()) and
                (pos + 3 >= len(expr) or not expr[pos + 3].isalnum())):
            return True

        # 检查 'or'
        if (pos + 2 <= len(expr) and
                expr[pos:pos + 2].lower() == 'or' and
                (pos == 0 or not expr[pos - 1].isalnum()) and
                (pos + 2 >= len(expr) or not expr[pos + 2].isalnum())):
            return True

        return False

    def _get_logical_operator_length(self, expr: str, pos: int) -> int:
        """获取逻辑运算符的长度"""
        if pos + 3 <= len(expr) and expr[pos:pos + 3].lower() == 'and':
            return 3
        elif pos + 2 <= len(expr) and expr[pos:pos + 2].lower() == 'or':
            return 2
        return 0

    def _split_expression_by_logical_operators(self, expression: str) -> List[str]:
        """
        已弃用：按逻辑运算符分割表达式的方法有括号平衡问题
        使用新的安全解析器替代此方法
        """
        logger.warning("_split_expression_by_logical_operators已弃用，存在括号平衡问题，使用安全解析器")

        # 使用安全解析器提取参数，然后转换为简单字符串列表以保持兼容性
        try:
            params = self.new_extract_param_details_fixed(expression)
            # 将参数信息转换回表达式字符串以保持向后兼容
            result = []
            for param in params:
                param_expr = f"{param['param_name']}{param.get('operator', '=')}{param.get('expected_value', '')}"
                result.append(param_expr)
            return result
        except Exception as e:
            logger.error(f"安全解析器失败: {str(e)}")
            return [expression.strip()]

    def _find_logical_operators_ignoring_param_brackets(self, expression: str) -> List[Tuple[int, int]]:
        """
        找到表达式中的逻辑运算符位置，忽略参数名中的括号
        返回: [(position, operator_length), ...]
        """
        logical_positions = []
        i = 0
        bracket_depth = 0  # 结构性括号深度

        while i < len(expression):
            char = expression[i]

            # 跟踪结构性括号深度
            if char == '(':
                # 检查是否是结构性左括号
                if self._is_structural_left_bracket(expression, i):
                    bracket_depth += 1
            elif char == ')' and bracket_depth > 0:
                # 检查是否是结构性右括号
                if self._is_structural_right_bracket(expression, i, bracket_depth):
                    bracket_depth -= 1

            # 只在结构性括号外面查找逻辑运算符
            if bracket_depth == 0:
                # 检查 'and'
                if (i + 3 <= len(expression) and
                        expression[i:i + 3].lower() == 'and' and
                        (i == 0 or not expression[i - 1].isalnum()) and
                        (i + 3 >= len(expression) or not expression[i + 3].isalnum())):
                    logical_positions.append((i, 3))
                    i += 3
                    continue

                # 检查 'or'
                elif (i + 2 <= len(expression) and
                      expression[i:i + 2].lower() == 'or' and
                      (i == 0 or not expression[i - 1].isalnum()) and
                      (i + 2 >= len(expression) or not expression[i + 2].isalnum())):
                    logical_positions.append((i, 2))
                    i += 2
                    continue

            i += 1

        return logical_positions

    def _is_structural_left_bracket(self, expression: str, pos: int) -> bool:
        """
        判断指定位置的左括号是否为结构性括号而非参数名中的括号
        结构性左括号的特征：
        1. 在表达式的开始
        2. 在逻辑运算符之后
        3. 在另一个结构性左括号之后
        """
        if pos == 0:
            return True

        # 查找前面的非空白字符
        prev_pos = pos - 1
        while prev_pos >= 0 and expression[prev_pos].isspace():
            prev_pos -= 1

        if prev_pos < 0:
            return True

        # 检查前面是否是逻辑运算符
        if prev_pos >= 2:
            if expression[prev_pos - 2:prev_pos + 1].lower() == 'and' or expression[
                prev_pos - 1:prev_pos + 1].lower() == 'or':
                return True
        if prev_pos >= 1:
            if expression[prev_pos - 1:prev_pos + 1].lower() == 'or':
                return True

        # 检查前面是否是另一个左括号
        if expression[prev_pos] == '(':
            return True

        return False

    def _is_structural_right_bracket(self, expression: str, pos: int, current_depth: int) -> bool:
        """
        判断指定位置的右括号是否为结构性括号
        通过匹配的左括号来判断
        """
        # 简化的判断：只要当前有结构性括号深度，就认为是结构性右括号
        return current_depth > 0

    def _parse_single_parameter_expression(self, expression: str) -> Optional[Dict[str, Any]]:
        """
        解析单个参数表达式，正确处理参数名中的括号
        例如：非同频测量RSRP触发门限(2dB)=12
        """
        if not expression or not expression.strip():
            return None

        expression = expression.strip()

        # 移除可能的外层结构性括号
        clean_expr = self._remove_outer_structural_brackets(expression)

        # 查找运算符，从最精确的开始（按长度降序）
        for operator in self.OPERATORS:
            # 从后向前查找运算符，避免参数名中的符号干扰
            op_positions = []
            start_search = 0
            while True:
                pos = clean_expr.find(operator, start_search)
                if pos == -1:
                    break
                op_positions.append(pos)
                start_search = pos + 1

            # 从最后一个可能的运算符开始尝试
            for pos in reversed(op_positions):
                param_name = clean_expr[:pos].strip()
                param_value = clean_expr[pos + len(operator):].strip()

                # 验证参数名和值是否有效
                if param_name and param_value and self._is_valid_parameter_split(param_name, param_value, operator):
                    return self._parse_param_detail(param_name, param_value, operator)

        # 没有找到有效的运算符
        logger.warning(f"无法解析参数表达式: {expression}")
        return None

    def _is_valid_parameter_split(self, param_name: str, param_value: str, operator: str) -> bool:
        """
        验证参数名和值的分割是否有效
        避免在参数名中间拆分或在不应该分割的地方分割
        """
        # 基本验证：参数名和值都不能为空
        if not param_name.strip() or not param_value.strip():
            return False

        # 验证参数名不应该以逻辑运算符结尾
        param_name_lower = param_name.lower().strip()
        if param_name_lower.endswith(' and') or param_name_lower.endswith(' or'):
            return False

        # 验证参数值不应该以逻辑运算符开始
        param_value_lower = param_value.lower().strip()
        if param_value_lower.startswith('and ') or param_value_lower.startswith('or '):
            return False

        # 验证括号匹配：参数名中的括号应该匹配
        open_brackets = param_name.count('(')
        close_brackets = param_name.count(')')
        if open_brackets != close_brackets:
            return False

        return True

    def _parse_simple_expression(self, expression: str) -> Dict[str, Any]:
        """解析简单表达式（逗号分隔的参数列表）"""
        expected_params = []
        # 按逗号分割多个参数表达式
        param_expressions = [expr.strip() for expr in expression.split(',') if expr.strip()]

        for param_expr in param_expressions:
            param_detail = self._parse_single_param_expr(param_expr)
            if param_detail:
                expected_params.append(param_detail)

        return {'type': 'simple', 'params': expected_params}

    def _parse_single_param_expr(self, param_expr: str) -> Optional[Dict[str, Any]]:
        """解析单个参数表达式（如"param>value"）"""
        for op in self.OPERATORS:
            op_index = param_expr.find(op)
            if op_index != -1:
                param_name = param_expr[:op_index].strip()
                param_value = param_expr[op_index + len(op):].strip()

                if not param_name:  # 无效的参数名
                    return None

                return self._parse_param_detail(param_name, param_value, op)

        return None  # 未找到有效运算符

    def _parse_param_detail(self, param_name: str, param_value: str, operator: str) -> Dict[str, Any]:
        """
        解析参数详细信息，统一处理单值和多值参数
        多值参数判断：包含&和:的组合
        """
        # 判断是否为多值参数
        if '&' in param_value and ':' in param_value:
            switches = {}
            for switch_expr in param_value.split('&'):
                if ':' in switch_expr:
                    switch_name, switch_state = switch_expr.split(':', 1)
                    switches[switch_name.strip()] = switch_state.strip()

            return {
                'param_name': param_name,
                'param_type': 'multiple',
                'operator': operator,
                'expected_switches': switches,
                'expected_value': param_value
            }
        else:
            # 单值参数
            return {
                'param_name': param_name,
                'param_type': 'single',
                'operator': operator,
                'expected_value': param_value
            }

    def parse_expected_expression(self, expression: str) -> Dict[str, Any]:
        """
        解析期望值表达式的入口方法 - 修正版

        支持格式:
        1. 简单格式:
           - 参数名1=值1,参数名2>值2,参数名3<=值3
           - 参数名1=k1:开&k2:关&k3:开
        2. 复杂格式:
           - （参数名1=值1 and 参数名2=值2）or（参数名3>值3 and 参数名2!=值2）

        Args:
            expression: 待解析的表达式字符串

        Returns:
            解析结果字典，包含类型和参数信息
        """
        if not expression or expression == 'nan':
            return {'type': 'simple', 'params': []}

        # 先进行标准化处理，解决"7and异频切换测量参数组标识=5"这类问题
        normalized_expression = self._normalize_condition_expression(expression)

        # 使用更准确的复杂表达式检测
        if self._is_complex_expression_improved(normalized_expression):
            return self._parse_complex_expression(normalized_expression)
        else:
            return self._parse_simple_expression(normalized_expression)

        # 此处的重复代码已移除，统一使用上面的_is_complex_expression和_parse_complex_expression方法

        # 简单格式：按逗号分割多个参数
        expected_params = []
        param_expressions = [expr.strip() for expr in expression.split(',') if expr.strip()]

        for param_expr in param_expressions:
            if '=' in param_expr:
                param_name, param_value = param_expr.split('=', 1)
                param_name = param_name.strip()
                param_value = param_value.strip()

                # 检查是否是多值参数（包含开关组合）
                if '&' in param_value and ':' in param_value:
                    # 多值参数
                    switches = {}
                    for switch_expr in param_value.split('&'):
                        if ':' in switch_expr:
                            switch_name, switch_state = switch_expr.split(':', 1)
                            switches[switch_name.strip()] = switch_state.strip()

                    expected_params.append({
                        'param_name': param_name,
                        'param_type': 'multiple',
                        'expected_switches': switches,
                        'expected_value': param_value
                    })
                else:
                    # 单值参数
                    expected_params.append({
                        'param_name': param_name,
                        'param_type': 'single',
                        'expected_value': param_value
                    })

        return {'type': 'simple', 'params': expected_params}

    # def validate_complex_expected_expression(self, expression: str, data_row: Dict[str, Any]) -> bool:
    #     """
    #     验证复杂的期望值表达式
    #     支持格式: （参数名1=值1and参数名2=值2）or（参数名3>值3and参数名2!=值2）
    #     """
    #     try:
    #         # 重用条件表达式解析的逻辑
    #         normalized_expr = self._normalize_condition_expression(expression)
    #
    #         def evaluate_single_validation(cond: str) -> bool:
    #             """评估单个验证条件"""
    #             cond = cond.strip()
    #
    #             # 支持的运算符
    #             operators = ['>=', '<=', '!=', '>', '<', '=']
    #
    #             for op in operators:
    #                 if op in cond:
    #                     parts = cond.split(op, 1)
    #                     if len(parts) == 2:
    #                         param_name = parts[0].strip()
    #                         expected_value = parts[1].strip()
    #
    #                         # 获取实际值
    #                         actual_value = str(data_row.get(param_name, '')).strip()
    #
    #                         # 对于多值参数，需要特殊处理
    #                         if '&' in expected_value and ':' in expected_value:
    #                             # 多值参数验证
    #                             return self._validate_multi_value_parameter(actual_value, expected_value)
    #                         else:
    #                             # 单值参数验证
    #                             return self._compare_values(actual_value, op, expected_value)
    #
    #             return False
    #
    #         # 递归处理括号和逻辑运算符
    #         def evaluate_expression(expr: str) -> bool:
    #             expr = expr.strip()
    #
    #             # 处理最内层括号
    #             while '(' in expr and ')' in expr:
    #                 start = expr.rfind('(')
    #                 end = expr.find(')', start)
    #                 if end == -1:
    #                     break
    #
    #                 # 评估括号内的表达式
    #                 inner_expr = expr[start + 1:end]
    #                 result = self._evaluate_simple_expression(inner_expr, evaluate_single_validation)
    #
    #                 # 替换括号及其内容为结果
    #                 expr = expr[:start] + str(result).lower() + expr[end + 1:]
    #
    #             # 评估剩余表达式
    #             return self._evaluate_simple_expression(expr, evaluate_single_validation)
    #
    #         return evaluate_expression(normalized_expr)
    #
    #     except Exception as e:
    #         logger.error(f"验证复杂期望值表达式失败: {expression}, 错误: {str(e)}")
    #         return False
    def _remove_outer_structural_brackets(self, expression: str) -> str:
        """
        移除最外层的结构性括号 - 安全版本
        注意：此方法现在使用安全的括号检查逻辑，避免递归过度剥离
        """
        expr = expression.strip()

        # 如果表达式不以括号开始和结束，直接返回
        if not (expr.startswith('(') and expr.endswith(')')):
            return expr

        # 使用安全的完整包围检查，避免过度剥离
        if self.is_fully_wrapped_by_brackets(expr):
            # 只去除一层外层括号，避免破坏内部结构
            return expr[1:-1].strip()
        else:
            return expr

    def validate_complex_expected_expression(self, expression: str, data_row: Dict[str, Any]) -> bool:
        """
        验证复杂的期望值表达式 - 使用新的智能解析逻辑
        支持参数名/值包含括号的情况，正确区分结构性括号和参数内容括号
        """
        try:
            # 使用新的解析逻辑来分析表达式结构
            parsed_result = self._parse_complex_expression(expression)

            if 'error' in parsed_result:
                logger.warning(f"解析表达式时出现错误: {parsed_result['error']}")
                return False

            params = parsed_result.get('params', [])
            if not params:
                logger.warning(f"未能从表达式中提取到有效参数: {expression}")
                return False

            # 评估每个参数条件
            param_results = []
            for param_info in params:
                param_name = param_info['param_name']
                operator = param_info.get('operator', '=')

                if param_name not in data_row:
                    logger.debug(f"数据中未找到参数: {param_name}")
                    param_results.append(False)
                    continue

                actual_value = str(data_row[param_name]).strip()

                if param_info['param_type'] == 'multiple':
                    # 多值参数验证
                    expected_switches = param_info['expected_switches']
                    is_match, _ = self._check_multi_value_match(actual_value, expected_switches)
                    param_results.append(is_match)
                else:
                    # 单值参数验证
                    expected_value = param_info['expected_value']
                    result = self._compare_values(actual_value, operator, expected_value)
                    param_results.append(result)

            # 根据原始表达式的逻辑结构计算最终结果
            return self._evaluate_parameter_results_with_logic(expression, param_results)

        except Exception as e:
            logger.error(f"验证复杂期望值表达式失败: {expression}, 错误: {str(e)}")
            return False

    def _evaluate_parameter_results_with_logic(self, expression: str, param_results: List[bool]) -> bool:
        """
        根据原始表达式的逻辑结构评估参数结果
        """
        if not param_results:
            return False

        if len(param_results) == 1:
            return param_results[0]

        # 简化的逻辑评估：根据表达式中的逻辑运算符来组合结果
        expr_lower = expression.lower()

        # 如果包含 'or'，任一为True则为True
        if ' or ' in expr_lower:
            return any(param_results)

        # 如果包含 'and' 或没有逻辑运算符，所有都必须为True
        return all(param_results)

    def _validate_multi_value_parameter(self, actual_value: str, expected_value: str) -> bool:
        """验证多值参数"""
        if not actual_value or not expected_value:
            return False

        # 解析期望的开关状态
        expected_switches = {}
        for switch_expr in expected_value.split('&'):
            if ':' in switch_expr:
                switch_name, switch_state = switch_expr.split(':', 1)
                expected_switches[switch_name.strip()] = switch_state.strip()

        # 使用现有的多值匹配逻辑
        is_match, _ = self._check_multi_value_match(actual_value, expected_switches)
        return is_match

    def execute_validation_rule(self, rule_id: str, data_groups: Dict[str, pd.DataFrame], sector_id, 
                               rule_chain: Optional[List[str]] = None) -> List[Dict[str, Any]]:
        """执行单个验证规则 - 修正版：支持数据筛选和传递，增加规则链跟踪"""
        if rule_id not in self.validation_rules:
            logger.warning(f"验证规则 {rule_id} 不存在")
            return []

        # 初始化或扩展规则链
        if rule_chain is None:
            rule_chain = []
        current_rule_chain = rule_chain + [rule_id]
        
        rule = self.validation_rules[rule_id]
        errors = []

        logger.info(f"执行验证规则: {rule_id} ({rule['check_type']}), 规则链: {' -> '.join(current_rule_chain)}")

        # 深拷贝数据，避免修改原始数据
        import copy
        working_data_groups = copy.deepcopy(data_groups)

        # 根据校验类型执行不同的验证，并获取通过验证的数据
        if rule['check_type'] == '漏配':
            check_errors, passed_data = self._check_missing_config(rule, working_data_groups, sector_id, current_rule_chain)
            errors.extend(check_errors)
            # 漏配检查后，传递找到期望配置的数据给后续验证
            filtered_data_groups = passed_data
        elif rule['check_type'] == '错配':
            check_errors, passed_data = self._check_incorrect_config(rule, working_data_groups, sector_id, current_rule_chain)
            errors.extend(check_errors)
            # 错配检查后，传递实际值与期望值匹配的数据给后续验证
            filtered_data_groups = passed_data
        else:
            logger.warning(f"未知的校验类型: {rule['check_type']}")
            filtered_data_groups = working_data_groups

        # 如果当前规则通过且有继续校验，使用筛选后的数据执行继续校验
        if not errors and rule['next_check_id']:
            logger.info(f"规则 {rule_id} 通过，继续执行: {rule['next_check_id']}")
            errors.extend(self.execute_validation_rule(rule['next_check_id'], filtered_data_groups, sector_id, current_rule_chain))
        elif errors:
            logger.info(f"规则 {rule_id} 检查失败，不继续后续验证")

        return errors

    def _filter_data_by_validation_result(self, data_groups: Dict[str, pd.DataFrame], rule: Dict[str, Any],
                                          passed_data: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """根据验证结果筛选数据，返回通过验证的数据子集"""
        # 返回通过验证的数据（去除了错配/漏配的数据）
        return passed_data

    def _check_missing_config(self, rule: Dict[str, Any], data_groups: Dict[str, pd.DataFrame], sector_id, 
                             rule_chain: Optional[List[str]] = None) -> Tuple[List[Dict[str, Any]], Dict[str, pd.DataFrame]]:
        """检查漏配"""
        mo_name = rule['mo_name']
        condition_expr = rule['condition_expression']
        expected_expr = rule['expected_expression']

        errors = []

        if mo_name not in data_groups:
            errors.append({
                'sector_id': "",
                'rule_id': rule['rule_id'],
                'mo_name': mo_name,
                'check_type': '漏配',
                'error_type': '数据不存在',
                'message': f'{mo_name}数据不存在',
                'error_description': rule['error_description']
            })
            return errors, {}

        mo_data = pd.DataFrame(data_groups[mo_name])
        expected_result = self.parse_expected_expression(expected_expr)

        if expected_result['type'] == 'simple' and not expected_result['params']:
            logger.warning(f"规则 {rule['rule_id']} 没有有效的期望值表达式")
            return errors, {}

        # 筛选出满足条件且满足期望的数据行
        validated_data_groups = {}
        validated_rows = []

        # 遍历所有数据行，筛选出符合条件且满足期望的行
        condition_matched_rows = []
        expected_matched_rows = []

        for idx, row in mo_data.iterrows():
            row_dict = row.to_dict()
            row_dict = {k: str(v).strip() for k, v in row_dict.items()}

            # 检查条件表达式
            if not self.parse_condition_expression(condition_expr, row_dict):
                continue

            condition_matched_rows.append((idx, row, row_dict))

            # 检查期望值
            row_meets_expectation = False

            if expected_result['type'] == 'complex':
                # 复杂表达式验证
                if self.validate_complex_expected_expression(expected_result['expression'], row_dict):
                    row_meets_expectation = True
            else:
                # 简单表达式验证
                all_params_match = True
                for expected_param in expected_result['params']:
                    param_name = expected_param['param_name']

                    if param_name not in row_dict:
                        all_params_match = False
                        break

                    if expected_param['param_type'] == 'multiple':
                        # 多值参数检查
                        actual_value = row_dict[param_name]
                        is_match, wrong_switches = self._check_multi_value_match(actual_value,
                                                                                 expected_param['expected_switches'])
                        if not is_match:
                            all_params_match = False
                            break
                    else:
                        # 单值参数检查
                        if row_dict[param_name] != expected_param['expected_value']:
                            all_params_match = False
                            break

                if all_params_match:
                    row_meets_expectation = True

            if row_meets_expectation:
                expected_matched_rows.append((idx, row, row_dict))
                validated_rows.append(row)

        # 如果没有找到满足期望的行，报告漏配
        if not expected_matched_rows:
            # 获取参数详细信息列表
            param_details = []
            param_names = []
            
            for param in expected_result['params']:
                param_name = param['param_name']
                param_names.append(param_name)
                param_info = self._get_parameter_info(mo_name, param_name)
                
                param_detail = {
                    'param_name': param_name,
                    'expected_value': param.get('expected_value', ''),
                    'parameter_description': param_info['parameter_description'],
                    'value_description': param_info['value_description'],
                    'parameter_type': param_info['parameter_type']
                }
                
                # 对于多值参数，添加开关信息
                if param.get('param_type') == 'multiple' and param.get('expected_switches'):
                    switch_descriptions = self._parse_value_descriptions(param_info['value_description'])
                    param_detail['expected_switches'] = param['expected_switches']
                    param_detail['switch_descriptions'] = switch_descriptions
                    
                param_details.append(param_detail)

            base_error = {
                'sector_id': sector_id,
                'rule_id': rule['rule_id'],
                'mo_name': mo_name,
                'param_names': param_names,
                'check_type': '漏配',
                'error_type': '漏配',
                'message': f'未找到符合条件的配置记录',
                'condition': condition_expr,
                'expected_expression': expected_expr,
                'error_description': rule['error_description'],
                'matched_condition_rows': len(condition_matched_rows),  # 满足条件但不满足期望的行数
                'total_rows': len(mo_data)  # 总数据行数
            }
            
            # 创建增强的错误记录
            enhanced_error = self._create_enhanced_error_record(base_error, rule, param_details, rule_chain)
            errors.append(enhanced_error)

        # 构建返回的数据字典：只包含满足期望表达式的数据行
        if validated_rows:
            validated_data_groups[mo_name] = pd.DataFrame(validated_rows)

        # 保留其他MO的数据不变
        for other_mo, other_data in data_groups.items():
            if other_mo != mo_name:
                validated_data_groups[other_mo] = other_data

        return errors, validated_data_groups

    def _check_incorrect_config(self, rule: Dict[str, Any], data_groups: Dict[str, pd.DataFrame], sector_id, 
                               rule_chain: Optional[List[str]] = None) -> Tuple[List[Dict[str, Any]], Dict[str, pd.DataFrame]]:
        """检查错配"""
        mo_name = rule['mo_name']
        condition_expr = rule['condition_expression']
        expected_expr = rule['expected_expression']

        errors = []

        if mo_name not in data_groups:
            errors.append({
                'sector_id': "",
                'rule_id': rule['rule_id'],
                'mo_name': mo_name,
                'check_type': '错配',
                'error_type': '数据不存在',
                'message': f'{mo_name}数据不存在',
                'error_description': rule['error_description']
            })
            return errors, {}  # 数据不存在时不报错配错误

        mo_data = pd.DataFrame(data_groups[mo_name])
        expected_result = self.parse_expected_expression(expected_expr)

        if expected_result['type'] == 'simple' and not expected_result['params']:
            logger.warning(f"规则 {rule['rule_id']} 没有有效的期望值表达式")
            return errors, {}

        # 筛选出满足条件且满足期望的数据行（正确配置的数据）
        validated_data_groups = {}
        validated_rows = []

        # 遍历所有数据行，检查错配并筛选正确配置
        for idx, row in mo_data.iterrows():
            row_dict = row.to_dict()
            row_dict = {k: str(v).strip() for k, v in row_dict.items()}

            # 检查条件表达式
            if not self.parse_condition_expression(condition_expr, row_dict):
                continue

            # 检查期望值，确定是否满足期望
            row_meets_expectation = False
            incorrect_params = []

            if expected_result['type'] == 'complex':
                # 复杂表达式验证
                if self.validate_complex_expected_expression(expected_result['expression'], row_dict):
                    row_meets_expectation = True
                else:
                    # 复杂表达式不满足，检查具体哪个参数错误
                    for expected_param in expected_result['params']:
                        param_name = expected_param['param_name']
                        if param_name not in row_dict:
                            continue

                        if expected_param['param_type'] == 'multiple':
                            actual_value = row_dict[param_name]
                            is_match, wrong_switches = self._check_multi_value_match(actual_value, expected_param[
                                'expected_switches'])
                            if not is_match:
                                incorrect_params.append((expected_param, actual_value, wrong_switches))
                        else:
                            if row_dict[param_name] != expected_param['expected_value']:
                                incorrect_params.append((expected_param, row_dict[param_name], None))
            else:
                # 简单表达式验证
                all_params_match = True
                for expected_param in expected_result['params']:
                    param_name = expected_param['param_name']
                    if param_name not in row_dict:
                        continue

                    if expected_param['param_type'] == 'multiple':
                        actual_value = row_dict[param_name]
                        is_match, wrong_switches = self._check_multi_value_match(actual_value,
                                                                                 expected_param['expected_switches'])
                        if not is_match:
                            all_params_match = False
                            incorrect_params.append((expected_param, actual_value, wrong_switches))
                    else:
                        if row_dict[param_name] != expected_param['expected_value']:
                            all_params_match = False
                            incorrect_params.append((expected_param, row_dict[param_name], None))

                if all_params_match:
                    row_meets_expectation = True

            # 如果行不满足期望，报告错配
            if not row_meets_expectation and incorrect_params:
                for expected_param, actual_value, wrong_switches in incorrect_params:
                    param_name = expected_param['param_name']

                    if expected_param['param_type'] == 'multiple':
                        # 多值参数错配
                        param_info = self._get_parameter_info(mo_name, param_name)
                        switch_descriptions = self._parse_value_descriptions(param_info['value_description'])

                        error_switch_descriptions = []
                        for wrong_switch in wrong_switches:
                            switch_name = wrong_switch['switch_name']
                            if switch_name in switch_descriptions:
                                error_switch_descriptions.append(f"{switch_name}: {switch_descriptions[switch_name]}")

                        base_error = {
                            'sector_id': row_dict.get('f_site_id', "") + "_" + row_dict.get('f_cell_id', ""),
                            'rule_id': rule['rule_id'],
                            'mo_name': mo_name,
                            'param_name': param_name,
                            'check_type': '错配',
                            'error_type': '错配',
                            'message': f'{param_name}开关配置错误',
                            'current_value': actual_value,
                            'expected_value': expected_param['expected_value'],
                            'wrong_switches': wrong_switches,
                            'switch_descriptions': error_switch_descriptions,
                            'condition': condition_expr,
                            'error_description': rule['error_description'],
                            'row_index': idx
                        }
                        
                        # 创建增强的错误记录
                        param_details = [{
                            'param_name': param_name,
                            'expected_value': expected_param['expected_value'],
                            'current_value': actual_value,
                            'parameter_description': param_info['parameter_description'],
                            'value_description': param_info['value_description'],
                            'parameter_type': param_info['parameter_type'],
                            'wrong_switches': wrong_switches
                        }]
                        
                        enhanced_error = self._create_enhanced_error_record(base_error, rule, param_details, rule_chain)
                        errors.append(enhanced_error)
                    else:
                        # 单值参数错配
                        param_info = self._get_parameter_info(mo_name, param_name)
                        
                        base_error = {
                            'sector_id': row_dict.get('f_site_id', "") + "_" + row_dict.get('f_cell_id', ""),
                            'rule_id': rule['rule_id'],
                            'mo_name': mo_name,
                            'param_name': param_name,
                            'check_type': '错配',
                            'error_type': '错配',
                            'message': f'{param_name}配置错误',
                            'current_value': actual_value,
                            'expected_value': expected_param['expected_value'],
                            'condition': condition_expr,
                            'error_description': rule['error_description'],
                            'row_index': idx
                        }
                        
                        # 创建增强的错误记录
                        param_details = [{
                            'param_name': param_name,
                            'expected_value': expected_param['expected_value'],
                            'current_value': actual_value,
                            'parameter_description': param_info['parameter_description'],
                            'value_description': param_info['value_description'],
                            'parameter_type': param_info['parameter_type']
                        }]
                        
                        enhanced_error = self._create_enhanced_error_record(base_error, rule, param_details, rule_chain)
                        errors.append(enhanced_error)

            # 只有满足期望的行才加入验证通过的数据中
            if row_meets_expectation:
                validated_rows.append(row)

        # 构建返回的数据字典：只包含正确配置的数据行
        if validated_rows:
            validated_data_groups[mo_name] = pd.DataFrame(validated_rows)

        # 保留其他MO的数据不变
        for other_mo, other_data in data_groups.items():
            if other_mo != mo_name:
                validated_data_groups[other_mo] = other_data

        return errors, validated_data_groups

    def _check_multi_value_match(self, actual_value: str, expected_switches: Dict[str, str]) -> Tuple[
        bool, List[Dict[str, str]]]:
        """
        检查多值参数是否匹配
        返回: (是否匹配, 错误的开关列表)
        """
        if not actual_value or not expected_switches:
            return False, []

        # 解析实际值中的开关状态
        actual_switches = {}
        for switch_expr in actual_value.split('&'):
            if ':' in switch_expr:
                switch_name, switch_state = switch_expr.split(':', 1)
                actual_switches[switch_name.strip()] = switch_state.strip()

        # 检查每个期望的开关状态，收集错误的开关
        wrong_switches = []
        all_match = True

        for switch_name, expected_state in expected_switches.items():
            if switch_name not in actual_switches:
                wrong_switches.append({
                    'switch_name': switch_name,
                    'expected_state': expected_state,
                    'actual_state': '未配置',
                    'error_type': '缺失'
                })
                all_match = False
            elif actual_switches[switch_name] != expected_state:
                wrong_switches.append({
                    'switch_name': switch_name,
                    'expected_state': expected_state,
                    'actual_state': actual_switches[switch_name],
                    'error_type': '错误'
                })
                all_match = False

        return all_match, wrong_switches

    def _get_parameter_info(self, mo_name: str, param_name: str) -> Dict[str, str]:
        """获取参数的完整信息"""
        if mo_name in self.parameter_info and param_name in self.parameter_info[mo_name]['parameters']:
            param_info = self.parameter_info[mo_name]['parameters'][param_name]
            return {
                'parameter_id': param_info.get('parameter_id', ''),
                'parameter_type': param_info.get('parameter_type', ''),
                'parameter_description': param_info.get('parameter_description', ''),
                'value_description': param_info.get('value_description', '')
            }
        return {
            'parameter_id': '',
            'parameter_type': '',
            'parameter_description': '',
            'value_description': ''
        }

    def _get_parameter_value_description(self, mo_name: str, param_name: str) -> str:
        """获取参数的值描述"""
        param_info = self._get_parameter_info(mo_name, param_name)
        return param_info['value_description']
        
    def _get_mo_description(self, mo_name: str) -> str:
        """获取MO对象描述"""
        if mo_name in self.parameter_info:
            return self.parameter_info[mo_name].get('mo_description', '')
        return ''
        
    def _create_enhanced_error_record(self, base_error: Dict[str, Any], rule: Dict[str, Any], 
                                    param_details: Optional[List[Dict[str, str]]] = None, 
                                    rule_chain: Optional[List[str]] = None) -> Dict[str, Any]:
        """创建增强的错误记录，包含参数含义和规则关系"""
        enhanced_error = base_error.copy()
        
        # 添加MO描述
        mo_name = enhanced_error.get('mo_name', '')
        enhanced_error['mo_description'] = self._get_mo_description(mo_name)
        
        # 添加参数详细信息
        if param_details:
            enhanced_error['parameter_details'] = param_details
        elif 'param_name' in enhanced_error:
            param_name = enhanced_error['param_name']
            param_info = self._get_parameter_info(mo_name, param_name)
            enhanced_error['parameter_info'] = param_info
            
        # 添加规则关系链
        if rule_chain:
            enhanced_error['rule_chain'] = rule_chain
            
        # 添加当前规则的完整信息
        enhanced_error['rule_info'] = {
            'rule_id': rule['rule_id'],
            'check_type': rule['check_type'],
            'condition_expression': rule['condition_expression'],
            'expected_expression': rule['expected_expression'],
            'next_check_id': rule.get('next_check_id', ''),
            'error_description': rule['error_description']
        }
        
        return enhanced_error

    def _parse_value_descriptions(self, value_description: str) -> Dict[str, str]:
        """
        解析值描述字符串，提取各个开关的说明
        格式: "beam1:第一波束开关,beam2:第二波束开关,beam3:第三波束开关"
        """
        descriptions = {}
        if not value_description:
            return descriptions

        for desc_part in value_description.split(','):
            if ':' in desc_part:
                switch_name, switch_desc = desc_part.split(':', 1)
                descriptions[switch_name.strip()] = switch_desc.strip()

        return descriptions

    def execute_validation_rule_with_tracking(self, rule_id: str, data_groups: Dict[str, pd.DataFrame], 
                                             sector_id, rule_chain: Optional[List[str]] = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """执行验证规则并追踪规则链"""
        if rule_chain is None:
            rule_chain = []
        
        current_chain = rule_chain + [rule_id]
        all_rule_chains = []
        
        # 执行规则并收集规则链信息
        errors = self._execute_single_rule_with_tracking(rule_id, data_groups, sector_id, current_chain, all_rule_chains)
        
        return errors, all_rule_chains
    
    def _execute_single_rule_with_tracking(self, rule_id: str, data_groups: Dict[str, pd.DataFrame], 
                                          sector_id, current_chain: List[str], 
                                          all_rule_chains: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """执行单个规则并记录规则链信息"""
        rule = self.validation_rules[rule_id]
        logger.info(f"执行验证规则: {rule_id} ({rule['check_type']}), 规则链: {' -> '.join(current_chain)}")
        
        # 记录规则链信息
        chain_info = {
            'chain': current_chain.copy(),
            'rule_id': rule_id,
            'rule_type': rule['check_type'],
            'description': rule['error_description'],
            'status': 'executed',
            'has_errors': False
        }
        
        # 执行具体的验证逻辑
        if rule['check_type'] == '漏配':
            errors, validated_data_groups = self._check_missing_config(rule, data_groups, sector_id, current_chain)
        elif rule['check_type'] == '错配':
            errors, validated_data_groups = self._check_incorrect_config(rule, data_groups, sector_id, current_chain)
        else:
            errors, validated_data_groups = [], data_groups
        
        # 更新规则链状态
        if errors:
            chain_info['has_errors'] = True
            chain_info['error_count'] = len(errors)
        
        all_rule_chains.append(chain_info)
        
        # 如果没有发现错误且有下一个检查规则，继续执行
        next_check_id = rule.get('next_check_id', '')
        if not errors and next_check_id:
            logger.info(f"规则 {rule_id} 通过，继续执行: {next_check_id}")
            next_chain = current_chain + [next_check_id]
            next_errors = self._execute_single_rule_with_tracking(next_check_id, validated_data_groups, 
                                                                sector_id, next_chain, all_rule_chains)
            errors.extend(next_errors)
        else:
            if errors:
                logger.info(f"规则 {rule_id} 检查失败，不继续后续验证")
        
        return errors

    def _generate_rule_execution_summary(self, executed_rule_chains: List[Dict[str, Any]], 
                                       all_errors: List[Dict[str, Any]]) -> str:
        """生成规则执行总结"""
        summary_lines = []
        summary_lines.append("\n\ud83d\udcca 规则执行流程总结:")
        summary_lines.append("=" * 50)
        
        # 统计总体情况
        total_chains = len(executed_rule_chains)
        error_chains = len([chain for chain in executed_rule_chains if chain['has_errors']])
        total_errors = len(all_errors)
        
        summary_lines.append(f"📊 总体统计:")
        summary_lines.append(f"   • 执行规则链数: {total_chains}")
        summary_lines.append(f"   • 有问题的规则链: {error_chains}")
        summary_lines.append(f"   • 发现问题总数: {total_errors}")
        summary_lines.append("")
        
        # 详细规则链分析
        summary_lines.append(f"🔍 验证流程分析:")
        
        for i, chain in enumerate(executed_rule_chains, 1):
            chain_str = " -> ".join(chain['chain'])
            status = "❌ 有问题" if chain['has_errors'] else "✅ 通过"
            
            summary_lines.append(f"   {i}. {chain_str}")
            summary_lines.append(f"      状态: {status}")
            summary_lines.append(f"      类型: {chain['rule_type']}")
            
            if chain['has_errors']:
                summary_lines.append(f"      问题数: {chain.get('error_count', 0)}")
            
            summary_lines.append(f"      说明: {chain['description']}")
            summary_lines.append("")
        
        # 问题类型统计
        if all_errors:
            error_types = {}
            for error in all_errors:
                error_type = error.get('error_type', '未知')
                error_types[error_type] = error_types.get(error_type, 0) + 1
            
            summary_lines.append(f"📊 问题类型分布:")
            for error_type, count in error_types.items():
                summary_lines.append(f"   • {error_type}: {count} 个")
            summary_lines.append("")
        
        summary_lines.append("💡 建议:")
        if total_errors == 0:
            summary_lines.append("   • 所有验证规则都通过，配置正常！")
        else:
            summary_lines.append(f"   • 发现 {total_errors} 个配置问题，建议优先处理错配问题")
            summary_lines.append("   • 检查规则链中的前置条件是否满足")
        
        summary_lines.append("=" * 50)
        
        return "\n".join(summary_lines)

    def validate_sector_data(self, data_groups: Dict[str, pd.DataFrame], sector_id) -> List[Dict[str, Any]]:
        """验证扇区数据"""
        all_errors = []
        executed_rule_chains = []  # 记录所有执行的规则链

        # 找到所有入口验证规则（没有被其他规则引用的规则）
        referenced_rules = set()
        for rule in self.validation_rules.values():
            if rule['next_check_id']:
                referenced_rules.add(rule['next_check_id'])

        entry_rules = [rule_id for rule_id in self.validation_rules.keys()
                       if rule_id not in referenced_rules]

        logger.info(f"发现 {len(entry_rules)} 个入口验证规则: {entry_rules}")

        # 执行每个入口规则
        for rule_id in entry_rules:
            errors, rule_chains = self.execute_validation_rule_with_tracking(rule_id, data_groups, sector_id)
            all_errors.extend(errors)
            executed_rule_chains.extend(rule_chains)

        # 生成规则关系总结
        rule_summary = self._generate_rule_execution_summary(executed_rule_chains, all_errors)
        logger.info(rule_summary)

        return all_errors

    def create_sample_excel(self) -> None:
        """创建示例Excel文件"""
        logger.info("正在生成示例参数知识库...")

        wb = Workbook()

        # 删除默认工作表
        if wb.active:
            default_ws = wb.active
            wb.remove(default_ws)

        # 创建参数信息表
        self._create_parameter_info_sheet(wb)

        # 创建验证规则表
        self._create_validation_rules_sheet(wb)

        # 保存文件
        wb.save(self.knowledge_file)
        logger.info(f"示例参数知识库已生成: {self.knowledge_file}")

    def _create_parameter_info_sheet(self, wb: Workbook) -> None:
        """创建参数信息表"""
        ws = wb.create_sheet("参数信息")

        # 设置表头
        headers = ['MO名称', 'MO描述', '场景类型', '参数名称', '参数ID', '参数类型', '参数含义', '值描述']
        ws.append(headers)

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 添加示例数据
        sample_data = [
            # NRCELL参数
            ["NRCELL", "5G小区对象", "5G基础配置", "跟踪区码", "tac", "single", "标识小区所属的跟踪区", ""],
            ["NRCELL", "5G小区对象", "5G基础配置", "小区状态", "cellState", "single", "小区的激活状态", ""],

            # NRDUCELL参数
            ["NRDUCELL", "5G DU小区对象", "5G物理层配置", "小区半径(米)", "cellRadius", "single", "小区覆盖半径", ""],
            ["NRDUCELL", "5G DU小区对象", "5G物理层配置", "最大传输功率", "maxTxPower", "single", "小区最大发射功率",
             ""],

            # NRDUCELLBEAM参数（多值参数）
            ["NRDUCELLBEAM", "5G波束配置对象", "5G波束管理", "波束开关组合", "beamSwitchComb", "multiple",
             "波束开关状态组合", "beam1:第一波束开关,beam2:第二波束开关,beam3:第三波束开关"],

            # NRCELLFREQRELATION参数
            ["NRCELLFREQRELATION", "小区频率关系对象", "频率管理配置", "连接态频率优先级", "connectedFreqPriority",
             "single", "连接态下的频率优先级", ""],
        ]

        for row_data in sample_data:
            ws.append(row_data)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_validation_rules_sheet(self, wb: Workbook) -> None:
        """创建验证规则表"""
        ws = wb.create_sheet("验证规则")

        # 设置表头
        headers = ['校验ID', '校验类型', 'MO名称', '条件表达式', '期望值表达式', '错误描述', '继续校验ID']
        ws.append(headers)

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 添加示例验证规则
        sample_rules = [
            # 复杂嵌套验证链示例
            ["MISS_001", "漏配", "NRCELL", "", "跟踪区码=100", "缺少跟踪区码为100的小区配置", "ERROR_001"],
            ["ERROR_001", "错配", "NRDUCELL", "跟踪区码=100", "小区半径(米)=500",
             "跟踪区码为100的小区，半径应配置为500米", "ERROR_002"],
            ["ERROR_002", "错配", "NRDUCELL", "跟踪区码=100and小区半径(米)=500", "最大传输功率=43",
             "半径500米的小区，功率应为43dBm", "ERROR_003"],
            ["ERROR_003", "错配", "NRDUCELLBEAM", "跟踪区码=100", "波束开关组合=beam1:开&beam2:关&beam3:开",
             "跟踪区码100的小区，波束组合应为beam1开beam2关beam3开", "MISS_002"],
            ["MISS_002", "漏配", "NRCELLFREQRELATION", "跟踪区码=100", "连接态频率优先级=1",
             "缺少跟踪区码100小区的频率优先级配置", "ERROR_004"],
            ["ERROR_004", "错配", "NRCELL", "跟踪区码=100and连接态频率优先级=1", "小区状态=激活",
             "已配置频率优先级的小区状态应为激活", ""],

            # 复杂条件示例
            ["COMPLEX_001", "错配", "NRDUCELL", "(跟踪区码=200or跟踪区码=300)and小区状态=激活", "小区半径(米)=1000",
             "特殊跟踪区的激活小区半径应为1000米", ""],
            ["COMPLEX_002", "漏配", "NRCELL", "小区半径(米)>500and最大传输功率>=40", "小区状态=激活",
             "大半径高功率小区必须激活", ""],
        ]

        for row_data in sample_rules:
            ws.append(row_data)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def run_validation_example(self) -> None:
        """运行验证示例"""
        logger.info("🧪 开始验证示例测试...")

        # 创建测试数据
        # 创建测试数据
        datas = {
            "result": {
                "NRDUCELL": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR DU小区标识": "4",
                        "小区半径(米)": "4000"
                    }
                ],
                "NRCELLALGOSWITCH": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "异频切换算法开关": "基于覆盖的异频切换开关:开&基于频率优先级的异频切换开关:关&异频重定向开关:开&基于运营商专用优先级的异频切换开关:关&音视频异频切换配合开关:开&基于业务的异频切换开关:关&基于覆盖的异频盲切换开关:关&FR1到FR2频点轮询选择开关:关&基于上行干扰的异频切换开关:关&基于SSB SINR的异频切换开关:关&NSA基于上行干扰的异频切换开关:关&异频切换配合开关:关&基于能效的异频切换开关:关&基于MBS兴趣指示的异频切换开关:关&基于业务的异频盲切换开关:关"
                    }
                ],
                "NRCELLFREQRELATION": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "7783",
                        "连接态频率优先级": "2",
                        "小区重选优先级": "6",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "14"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "5361",
                        "连接态频率优先级": "1",
                        "小区重选优先级": "5",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "10"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "7714",
                        "连接态频率优先级": "2",
                        "小区重选优先级": "6",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "10"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "SSB频域位置": "7853",
                        "连接态频率优先级": "2",
                        "小区重选优先级": "6",
                        "最低接收电平(2dBm)": "-64",
                        "低优先级重选门限(2dB)": "10"
                    }
                ],
                "NRCELLINTERFHOMEAGRP": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "异频切换测量参数组标识": "0",
                        "基于覆盖的异频A5 RSRP触发门限1(dBm)": "-105",
                        "基于覆盖的异频A5 RSRP触发门限2(dBm)": "-100",
                        "基于覆盖的异频A2 RSRP触发门限(dBm)": "-105",
                        "基于覆盖的异频A1 RSRP触发门限(dBm)": "-100",
                        "异频测量事件时间迟滞(毫秒)": "320",
                        "异频测量事件幅度迟滞(0.5dB)": "2",
                        "异频A1A2时间迟滞(毫秒)": "320",
                        "异频A1A2幅度迟滞(0.5dB)": "2"
                    }
                ],
                "NRCELLQCIBEARER": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "1",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "2",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "3",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "4",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "5",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "6",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "7",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "8",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "9",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "65",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "66",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "69",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "70",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "75",
                        "异频切换测量参数组标识": "0"
                    },
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "服务质量等级": "79",
                        "异频切换测量参数组标识": "0"
                    }
                ],
                "NRCELLRESELCONFIG": [
                    {
                        "f_site_id": "13566583",
                        "f_cell_id": "4",
                        "Name": "YC5GHTA东台_城南OLT5G宏站BBU4_[13566583][130.1][场景]",
                        "f_local_cell_id": "4",
                        "NR小区标识": "4",
                        "非同频测量RSRP触发门限(2dB)": "10",
                        "服务频点低优先级RSRP重选门限(2dB)": "9"
                    }
                ]
            }
        }
        test_datas = datas['result']
        # 执行验证
        sector_datas = {}
        for mo, raw_data in test_datas.items():
            data_df = pd.DataFrame(raw_data)

            # 按站点和小区ID分组
            for (site_id, cell_id), sector_df in data_df.groupby(
                    ['f_site_id', 'f_cell_id'],
                    dropna=False
            ):
                sector_id = f"{site_id}_{cell_id}"

                # 如果sector_id不在外层字典中，初始化一个空字典
                if sector_id not in sector_datas:
                    sector_datas[sector_id] = {}

                # 将当前MO的数据放入对应的sector_id下
                sector_datas[sector_id][mo] = sector_df

        all_errors = []
        for sector_id, sector_dfs in sector_datas.items():
            errors = self.validate_sector_data(sector_dfs, sector_id)
            all_errors.extend(errors)

        # 输出结果
        if all_errors:
            logger.info(f"🔍 发现 {len(all_errors)} 个验证问题:")
            for i, error in enumerate(all_errors, 1):
                logger.info(f"   {i}. 【{error['check_type']}】{error.get('rule_id', 'N/A')} - {error['mo_name']}")
                if 'param_name' in error:
                    logger.info(f"      参数: {error['param_name']}")
                if 'param_names' in error:
                    logger.info(f"      参数: {', '.join(error['param_names'])}")
                logger.info(f"      错误: {error['message']}")

                # 处理多值参数的开关错误详情
                if 'wrong_switches' in error and error['wrong_switches']:
                    logger.info(f"      开关错误详情:")
                    for switch_error in error['wrong_switches']:
                        logger.info(
                            f"        - {switch_error['switch_name']}: 期望{switch_error['expected_state']}, 实际{switch_error['actual_state']}")

                    # 显示错误开关的描述
                    if 'switch_descriptions' in error and error['switch_descriptions']:
                        logger.info(f"      开关说明:")
                        for desc in error['switch_descriptions']:
                            logger.info(f"        - {desc}")

                # 显示单值参数的期望值和实际值
                elif 'current_value' in error and 'expected_value' in error:
                    logger.info(f"      期望值: {error['expected_value']}")
                    logger.info(f"      实际值: {error['current_value']}")

                if error.get('error_description'):
                    logger.info(f"      说明: {error['error_description']}")
                logger.info("")
        else:
            logger.info("✅ 所有验证规则都通过了")

    def new_extract_param_details_fixed(self, expression: str) -> List[Dict[str, Any]]:
        """
        完全重写的参数详细信息提取方法 - 正确处理括号平衡
        使用正确的递归下降解析，避免先剥离括号再分割导致的括号不匹配问题
        """
        try:
            logger.info(f"🔧 使用完全重写的解析器解析表达式: {expression}")

            # 预处理：仅标准化，不删除任何括号
            expr = self._normalize_condition_expression(expression).strip()

            # 核心逻辑：在保持括号完整的情况下查找主要逻辑运算符
            main_op = self.find_main_logical_op_safe(expr)

            if main_op is None:
                # 没有逻辑运算符，这是单个参数或被括号包围的表达式
                return self.parse_atomic_expression(expr)

            # 找到主要逻辑运算符，在该位置安全分割
            op_pos, op_len, op_type = main_op
            left_part = expr[:op_pos].strip()
            right_part = expr[op_pos + op_len:].strip()

            logger.info(f"安全分割为: '{left_part}' {op_type} '{right_part}'")

            # 递归解析两部分
            result = []
            result.extend(self.new_extract_param_details_fixed(left_part))
            result.extend(self.new_extract_param_details_fixed(right_part))

            logger.info(f"递归解析结果: {len(result)} 个参数")
            return result

        except Exception as e:
            logger.error(f"重写版解析器失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def is_fully_wrapped_by_brackets(self, expr: str) -> bool:
        """检查表达式是否完全被外层括号包围"""
        if not (expr.startswith('(') and expr.endswith(')')):
            return False

        bracket_count = 0
        for i, char in enumerate(expr):
            if char == '(':
                bracket_count += 1
            elif char == ')':
                bracket_count -= 1
                # 如果在结尾之前括号就归零了，说明不是完全包围
                if bracket_count == 0 and i < len(expr) - 1:
                    return False

        return bracket_count == 0

    def find_main_logical_op_safe(self, expr: str) -> Optional[Tuple[int, int, str]]:
        """
        安全查找主要逻辑运算符 - 严格遵守括号边界
        优先级：or > and（or优先级更低，应该先处理）
        """
        bracket_depth = 0

        # 第一遍：查找 'or'（优先级最低）
        for i in range(len(expr)):
            char = expr[i]

            if char == '(':
                bracket_depth += 1
            elif char == ')':
                bracket_depth -= 1
            elif bracket_depth == 0:  # 只在括号外查找
                # 检查是否为完整的 'or' 单词
                if (i + 2 <= len(expr) and
                        expr[i:i + 2].lower() == 'or' and
                        (i == 0 or not expr[i - 1].isalnum()) and
                        (i + 2 >= len(expr) or not expr[i + 2].isalnum())):
                    return (i, 2, 'or')

        # 第二遍：查找 'and'（优先级较高）
        bracket_depth = 0
        for i in range(len(expr)):
            char = expr[i]

            if char == '(':
                bracket_depth += 1
            elif char == ')':
                bracket_depth -= 1
            elif bracket_depth == 0:  # 只在括号外查找
                # 检查是否为完整的 'and' 单词
                if (i + 3 <= len(expr) and
                        expr[i:i + 3].lower() == 'and' and
                        (i == 0 or not expr[i - 1].isalnum()) and
                        (i + 3 >= len(expr) or not expr[i + 3].isalnum())):
                    return (i, 3, 'and')

        return None

    def find_main_logical_op(self, expr: str) -> Optional[Tuple[int, int, str]]:
        """旧版本，保持兼容性"""
        return self.find_main_logical_op_safe(expr)

    def parse_atomic_expression(self, expr: str) -> List[Dict[str, Any]]:
        """
        解析原子表达式（单个参数或被完整括号包围的表达式）
        确保括号平衡的前提下处理表达式
        """
        expr = expr.strip()
        if not expr:
            return []

        logger.info(f"解析原子表达式: {expr}")

        # 检查是否被完整的括号包围
        if expr.startswith('(') and expr.endswith(')') and self.is_fully_wrapped_by_brackets(expr):
            # 去除外层括号，递归解析内部表达式
            inner_expr = expr[1:-1].strip()
            logger.info(f"去除外层括号，递归解析: {inner_expr}")
            return self.new_extract_param_details_fixed(inner_expr)

        # 不是被括号包围的，应该是单个参数表达式
        param = self.parse_single_parameter(expr)
        return [param] if param else []

    def parse_single_parameter(self, expr: str) -> Optional[Dict[str, Any]]:
        """
        解析单个参数表达式 - 严格验证无逻辑运算符
        """
        expr = expr.strip()
        if not expr:
            return None

        logger.info(f"解析单个参数: {expr}")

        # 首先验证这个表达式中没有逻辑运算符（在括号外）
        if self.find_main_logical_op_safe(expr) is not None:
            logger.error(f"单个参数表达式中发现逻辑运算符: {expr}")
            return None

        # 寻找运算符（从最长的开始，避免嵌套问题）
        for operator in self.OPERATORS:
            # 从右往左查找，避免参数名中的符号干扰
            for i in range(len(expr) - len(operator), -1, -1):
                if expr[i:i + len(operator)] == operator:
                    param_name = expr[:i].strip()
                    param_value = expr[i + len(operator):].strip()

                    # 验证这是一个有效的分割
                    if param_name and param_value:
                        logger.info(f"找到参数: '{param_name}' {operator} '{param_value}'")
                        return self._parse_param_detail(param_name, param_value, operator)

        logger.warning(f"无法解析单个参数: {expr}")
        return None

    def parse_single_param_fixed(self, expr: str) -> Optional[Dict[str, Any]]:
        """旧版本，保持兼容性"""
        return self.parse_single_parameter(expr)


def main():
    """主程序入口"""
    try:
        logger.info("🚀 启动参数核查系统...")

        # 创建参数核查器实例
        checker = ParameterChecker()

        # 运行验证示例
        logger.info("🧪 开始验证示例测试...")
        checker.run_validation_example()

        logger.info("✨ 参数核查系统运行完成！")
        logger.info("📋 系统特性:")
        logger.info("   • 双分表设计：参数信息与验证规则完全分离")
        logger.info("   • 复杂条件支持：(param1=value1and param2=value2)or(param3>value3)")
        logger.info("   • 嵌套验证链：支持漏配↔错配无限嵌套调用")
        logger.info("   • 多值参数处理：beam1:开&beam2:关&beam3:开格式")
        logger.info("   • 智能条件筛选：先筛选符合条件的行再进行验证")

        return True

    except Exception as e:
        logger.error(f"程序运行出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


# These methods are now properly defined within the ParameterChecker class



if __name__ == "__main__":

    success = main()
    if not success:
        exit(1)
